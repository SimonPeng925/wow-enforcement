"""
WOW English 京东维权提取工具 v3.0
使用 Chrome 的真实登录状态（接管已登录浏览器）
"""

import os
import sys
import io
import time
import socket

# Windows 编码修复
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    try:
        os.system('chcp 65001 >nul 2>&1')
    except:
        pass

try:
    from playwright.sync_api import sync_playwright, Page
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import easyocr
except ImportError as e:
    print(f"[ERROR] Missing: {e}")
    print("Run: pip install playwright openpyxl easyocr")
    sys.exit(1)

# 预加载 OCR 读取器（中英文，GPU=False）
print("[OCR] 初始化文字识别引擎（首次运行需下载模型，约30秒）...")
_OCR_READER = None
def get_ocr_reader():
    global _OCR_READER
    if _OCR_READER is None:
        _OCR_READER = easyocr.Reader(['ch_sim', 'en'], gpu=False, verbose=False)
    return _OCR_READER


# ==================== 配置 ====================
BRAND_KEYWORDS = [
    "wow english", "wowenglish",
    "史蒂夫英语", "Steve English",
    "英语启蒙动画", "English Singsing",
    "wow english", "史蒂夫",
]
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")


# ==================== 京东提取器 ====================
class JDExtractor:
    """京东商品信息提取器 - v3"""

    # 京东页面选择器（新版 2024-2025）
    TITLE_SELECTORS = [
        # 新版页面（优先级高）
        '.product-title__main',
        '[class*="product-title"]',
        '[class*="sku-name"]',
        '[class*="item-name"]',
        '.p-title',
        '#itemName',
        # 旧版兼容
        'div[class*="product-name"] h1',
        '#name h1',
        'h1[class*="name"]',
        # 兜底
        'h1[class*="product"]',
        'h1[class*="item"]',
        'h1[class*="title"]',
        '[data-sku]',
        'h1',
    ]

    PRICE_SELECTORS = [
        # 新版页面
        '[class*="price"] [class*="integer"]',
        '[class*="price"] strong',
        '.p-price .price',
        '#jdprice',
        '[data-price]',
        '[class*="sale-price"]',
        '[class*="price-wrap"] span',
        # 京东 price.js 注入的数据
        '.J-p-',
        '#spec-qty-price',
    ]

    SHOP_SELECTORS = [
        # 店铺名称 - 最准确的选择器
        '[class*="shop-name"] a[href*="shop"]',
        '.j-shop-content .shop-name',
        '#shopInfo .shop-name',
        '#shop-name',
        '[class*="seller-info"] [class*="name"]',
        # 店铺浮层
        '#pop_shop .name',
        '.popshop .shop-name',
        # 旧版
        '[class*="shop-name"] a',
        '.shop-name a',
        '[class*="seller"] a',
        '[class*="store"] a',
        '[class*="shop"]',
    ]

    SELLER_ID_SELECTORS = [
        # 卖家ID
        '#pop_shop .name',
        '[class*="seller-code"]',
        '[class*="shop-id"]',
        '[data-sellerid]',
        '.shop-name',
        # 尝试从店铺链接提取
        '[class*="shop-name"] a[href*="shop.jd.com"]',
    ]

    SALES_SELECTORS = [
        # 销量/评价数
        '[class*="sale"]',
        '[class*="count"]',
        '.p-commit a',
        '#commit_count',
        '[class*="evaluate"]',
        '[class*="review"]',
        # 京东评价数字
        '.J-comm-count',
        '#count .number',
        '[class*="item"] [class*="count"]',
    ]

    MAIN_IMG_SELECTORS = [
        '#spec-img',
        '[class*="spec-img"] img',
        '[class*="main-img"] img',
        '[class*="preview"] img',
        '#large',
    ]

    @classmethod
    def extract(cls, page: Page, url: str) -> dict:
        data = {
            "platform": "京东",
            "title": "",
            "price": "",
            "shop_name": "",
            "seller_id": "",
            "sales": "",
            "main_images": [],
            "url": url,
            "infringement_check": "否"
        }

        # ── 先用 JS 直接从页面变量/JSON-LD 中提取最准确的数据 ──
        js_data = cls._extract_via_js(page)
        if js_data:
            data.update(js_data)
        else:
            # 回退到 CSS 选择器
            data["title"] = cls._try_selectors(page, cls.TITLE_SELECTORS)
            data["price"] = cls._try_selectors(page, cls.PRICE_SELECTORS)
            data["shop_name"] = cls._try_selectors(page, cls.SHOP_SELECTORS)
            data["seller_id"] = cls._try_selectors(page, cls.SELLER_ID_SELECTORS)
            data["sales"] = cls._try_selectors(page, cls.SALES_SELECTORS)

        data["main_images"] = cls._extract_main_images(page)
        data["infringement_check"] = cls._check_infringement(data)

        return data

    @staticmethod
    def _extract_via_js(page: Page) -> dict:
        """通过 JavaScript 直接读取京东页面注入的数据"""
        try:
            result = page.evaluate("""
                () => {
                    const data = {};
                    // 噪音词黑名单（遇到直接跳过）
                    const noiseList = ['最小单价', '搭配购买', '似乎出了点问题', '先看看', '极速审核',
                                       '立即购买', '加入购物车', '查看全部', '售后服务', '退换无忧'];

                    function isNoise(text) {
                        return noiseList.some(n => text.includes(n));
                    }

                    // 1. 商品标题 - 找主标题（大字，位于页面顶部）
                    const h1 = document.querySelector('h1') || document.querySelector('[class*="product-title"]');
                    if (h1) {
                        const t = h1.innerText.trim().substring(0, 200);
                        if (t && t.length > 5 && !isNoise(t)) data.title = t;
                    }

                    // 2. 价格 - 找数字开头的价格
                    const priceEls = document.querySelectorAll('[class*="price"]');
                    for (const el of priceEls) {
                        const t = el.innerText.trim();
                        if (/^￥|^¥|^\\d/.test(t) && !isNoise(t) && t.length < 30) {
                            data.price = t;
                            break;
                        }
                    }

                    // 3. 店铺名称 - 京东店铺信息栏（顶部评分区域）
                    const shopSelectors = [
                        // 新版京东店铺名（评分旁边）
                        '[class*="shop-info"] [class*="name"]',
                        '[class*="shop-header"] [class*="name"]',
                        '[class*="seller-info"] [class*="name"]',
                        // 店铺评分区域
                        '[class*="score"] [class*="shop"]',
                        '[class*="shop-name"]',
                        // 店铺链接
                        'a[href*="shop.jd.com"]',
                        '[class*="shop-name"] a',
                        '#shop-name a',
                        '#pop_shop a',
                    ];
                    for (const sel of shopSelectors) {
                        const el = document.querySelector(sel);
                        if (el) {
                            const t = el.innerText.trim();
                            if (t && !isNoise(t) && t.length > 1 && t.length < 50) {
                                data.shop_name = t;
                                break;
                            }
                        }
                    }
                    // 从店铺链接提取 seller_id
                    const shopLink = document.querySelector('a[href*="shop.jd.com"]') ||
                                     document.querySelector('[class*="shop-name"] a');
                    if (shopLink && shopLink.href) {
                        const m = shopLink.href.match(/shopId[=:]"?(\\d+)/) ||
                                 shopLink.href.match(/\\/shop\\/(\\d+)/) ||
                                 shopLink.href.match(/\\/(\\d+)\\.html.*shop/);
                        if (m) data.seller_id = m[1];
                    }

                    // 4. 销量/评价数
                    const salesEl = document.querySelector('.p-commit a') ||
                                    document.querySelector('#commit_count') ||
                                    document.querySelector('.J-comm-count');
                    if (salesEl) {
                        const t = salesEl.innerText.trim().split('\\n')[0];
                        if (t && !isNoise(t) && t.length < 30) data.sales = t;
                    }

                    return data;
                }
            """)
            # 只返回有值的字段
            return {k: v for k, v in result.items() if v}
        except Exception as e:
            return {}

    @staticmethod
    def _try_selectors(page: Page, selectors: list) -> str:
        for sel in selectors:
            try:
                el = page.locator(sel).first
                if el.count() > 0 and el.is_visible():
                    text = el.inner_text().strip()
                    if text and len(text) > 1:
                        return text[:200]
            except:
                continue
        return "未提取"

    @staticmethod
    def _extract_main_images(page: Page) -> list:
        images = []
        for sel in JDExtractor.MAIN_IMG_SELECTORS:
            try:
                els = page.locator(sel).all()
                for el in els[:5]:
                    if el.is_visible():
                        src = el.get_attribute('src') or el.get_attribute('data-src') or ''
                        if src and 'blank' not in src.lower():
                            import re
                            src = re.sub(r'/\d+x\d+/', '/800x800/', src)
                            images.append(src)
                if images:
                    break
            except:
                continue
        return list(dict.fromkeys(images))

    @staticmethod
    def _check_infringement(data: dict) -> str:
        text = f"{data.get('title', '')} {data.get('shop_name', '')}".lower()
        matched = [kw for kw in BRAND_KEYWORDS if kw.lower() in text]
        return f"是（匹配：{', '.join(matched)}）" if matched else "否"

    @staticmethod
    def _ocr_shop_name(image_path: str) -> str:
        """OCR 兜底：从截图顶部区域识别店铺名称"""
        try:
            from PIL import Image
            reader = get_ocr_reader()
            img = Image.open(image_path)
            w, h = img.size
            # 截取页面顶部 25% 的区域（店铺名通常在这里）
            top_region = img.crop((0, 0, w, int(h * 0.25)))
            top_path = image_path.replace('.png', '_shop_area.png')
            top_region.save(top_path)

            results = reader.readtext(top_path)
            # 合并所有识别文字
            all_text = ' '.join([r[1] for r in results])

            # 匹配常见店铺名模式：XX旗舰店/专营店/专卖店/直营店 等
            import re
            shop_match = re.search(
                r'([\u4e00-\u9fa5]{2,30}(?:旗舰店|专营店|专卖店|直营店|官方旗舰店|海外旗舰店|
                                    官方店|商城|百货|书屋|书店|
                                    \u300e[\u4e00-\u9fa5]{2,20}\u300f))',
                all_text
            )
            if shop_match:
                return shop_match.group(1).strip()
            # 没有匹配到店铺名模式，返回最长的中文字符串
            chinese_words = re.findall(r'[\u4e00-\u9fa5]{3,}', all_text)
            if chinese_words:
                return max(chinese_words, key=len)
        except Exception as e:
            print(f"[OCR WARN] OCR failed: {e}")
        return "未提取"


# ==================== 截图管理器 ====================
class ScreenshotManager:
    def __init__(self, page: Page, save_dir: str):
        self.page = page
        self.save_dir = save_dir
        os.makedirs(os.path.join(save_dir, "main_images"), exist_ok=True)

    def screenshot_page(self, full_page: bool = True) -> str:
        filename = os.path.join(self.save_dir, "screenshot.png")
        if full_page:
            h = self.page.evaluate("document.body.scrollHeight")
            self.page.set_viewport_size({"width": 1366, "height": min(h, 6000)})
        self.page.screenshot(path=filename, full_page=full_page)
        return filename

    def download_main_images(self, urls: list) -> list:
        saved = []
        for i, url in enumerate(urls[:5]):
            try:
                url = url.replace('_50x50', '').replace('_b.jpg', '.jpg')
                filename = os.path.join(self.save_dir, "main_images", f"jd_main_{i+1}.jpg")
                resp = self.page.context.request.get(url)
                if resp.ok:
                    with open(filename, 'wb') as f:
                        f.write(resp.body())
                    saved.append(filename)
            except Exception as e:
                pass
        return saved


# ==================== Excel 导出 ====================
class ExcelExporter:
    def __init__(self, data: dict, save_dir: str):
        self.data = data
        self.save_dir = save_dir

    def export(self) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "京东维权信息"

        header_fill = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        headers = ["平台", "商品标题", "商品链接", "价格", "店铺名称",
                   "卖家ID", "销量/评价", "主图URL", "提取时间", "疑似侵权"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")

        main_urls = "\n".join(self.data.get("main_images", []))
        row_data = [
            self.data.get("platform", ""),
            self.data.get("title", ""),
            self.data.get("url", ""),
            self.data.get("price", ""),
            self.data.get("shop_name", ""),
            self.data.get("seller_id", ""),
            self.data.get("sales", ""),
            main_urls,
            self.data.get("timestamp", ""),
            self.data.get("infringement_check", "")
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=2, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")

        widths = [10, 45, 60, 12, 25, 20, 15, 60, 20, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        if "是" in self.data.get("infringement_check", ""):
            ws.cell(row=2, column=10).fill = PatternFill(start_color="FF0000", fill_type="solid")
            ws.cell(row=2, column=10).font = Font(bold=True, color="FFFFFF")

        filename = os.path.join(self.save_dir, "jd_extraction.xlsx")
        wb.save(filename)
        return filename


# ==================== 举报模板 ====================
class ReportGenerator:
    def __init__(self, data: dict, save_dir: str):
        self.data = data
        self.save_dir = save_dir

    def generate(self) -> str:
        template = f"""
{'='*60}
               京东平台侵权举报信息
{'='*60}

【商品信息】
平台：{self.data.get('platform')}
商品标题：{self.data.get('title')}
商品链接：{self.data.get('url')}
当前价格：{self.data.get('price')}

【店铺信息】
店铺名称：{self.data.get('shop_name')}
卖家ID：{self.data.get('seller_id')}
销量/评价：{self.data.get('sales')}

【侵权分析】
关键词匹配：{self.data.get('infringement_check')}

【证据文件】
- 页面截图：screenshot.png
- 商品主图：main_images/
- 结构化数据：jd_extraction.xlsx

{'='*60}
提取时间：{self.data.get('timestamp')}
证据目录：{self.save_dir}
{'='*60}
"""
        filename = os.path.join(self.save_dir, "jd_report.txt")
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(template)
        return filename


# ==================== CDP 连接检测 ====================
def find_cdp_port():
    """检测 Chrome 调试端口"""
    for port in [9222, 9223, 9224]:
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1)
            result = sock.connect_ex(('127.0.0.1', port))
            sock.close()
            if result == 0:
                return port
        except:
            continue
    return None


# ==================== 主程序 ====================
def main():
    print("""
╔══════════════════════════════════════════════════════════╗
║      WOW English 京东维权提取工具 v3.0                   ║
║      🔗 支持直接使用已登录的 Chrome 浏览器               ║
╚══════════════════════════════════════════════════════════╝
    """)

    # 获取链接
    url = None
    if len(sys.argv) > 1:
        url = sys.argv[1]
    else:
        print("[INPUT] 请输入京东商品链接:")
        url = input(">>> ").strip()

    if not url or 'jd.com' not in url:
        print("[ERROR] 请提供京东商品链接")
        print("[USAGE] python jd_extract_v3.py <商品链接>")
        return

    # 创建输出目录
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_dir = os.path.join(OUTPUT_DIR, f"jd_{timestamp}")
    os.makedirs(save_dir, exist_ok=True)

    print(f"\n[1/4] URL: {url}")
    print(f"[2/4] Output: {save_dir}")

    # 检测 Chrome 调试端口
    cdp_port = find_cdp_port()
    if cdp_port:
        print(f"\n[CDP] 检测到 Chrome 调试端口 :{cdp_port}")
        print("[CDP] 尝试接管已打开的 Chrome...")
    else:
        print("""
[INFO] 未检测到 Chrome 调试模式！
      
如需使用已登录的 Chrome，请先：
  1. 关闭所有 Chrome 窗口
  2. 双击「启动已登录Chrome.bat」
  3. 登录京东后，保持浏览器开着，再运行本脚本

  或直接回车，我会启动独立浏览器（需要重新登录）
""")
        use_standalone = input("按回车启动独立浏览器，或输入 'q' 退出: ").strip()
        if use_standalone.lower() == 'q':
            return
        print("[BROWSER] 启动独立浏览器...")

    print("[3/4] Connecting to browser...")
    print("-" * 50)

    with sync_playwright() as p:
        browser = None
        page = None

        # ── 方案A：接管已打开的 Chrome ──
        if cdp_port:
            try:
                cdp_url = f"http://127.0.0.1:{cdp_port}"
                browser = p.chromium.connect_over_cdp(cdp_url)
                ctx = browser.contexts[0]

                # 查找已有的京东标签页
                jd_pages = [pg for pg in ctx.pages if 'jd.com' in pg.url and 'item.jd.com' in pg.url]
                if jd_pages:
                    page = jd_pages[0]
                    print(f"[CDP] 已接管京东商品页")
                else:
                    page = ctx.new_page()
                    print(f"[CDP] 已连接，在浏览器中新开标签页")
                    print(f"[CDP] 正在导航到商品页...")

                # 直接导航到目标 URL
                page.goto(url, wait_until="domcontentloaded", timeout=30000)
                print(f"[CDP] 页面已加载: {page.url[:60]}...")

            except Exception as e:
                print(f"[CDP] 接管失败: {e}")
                cdp_port = None  # 回退到方案B

        # ── 方案B：独立浏览器 ──
        if cdp_port is None or browser is None:
            try:
                browser = p.chromium.launch(
                    headless=False,
                    args=[
                        '--disable-blink-features=AutomationControlled',
                        '--disable-dev-shm-usage',
                        '--no-sandbox',
                    ]
                )
                ctx = browser.new_context(
                    viewport={"width": 1366, "height": 768},
                    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    locale='zh-CN',
                    no_viewport=True
                )
                page = ctx.new_page()
                page.goto(url, wait_until="domcontentloaded", timeout=30000)
                print("[BROWSER] 独立浏览器已启动")
            except Exception as e:
                print(f"[ERROR] 浏览器启动失败: {e}")
                return

        # ── 等待页面加载完成 ──
        print("[4/4] 等待页面加载...")
        # 等待网络空闲（京东有大量异步请求）
        try:
            page.wait_for_load_state("networkidle", timeout=15000)
        except:
            print("[WARN] 网络未完全空闲，继续...")
        page.wait_for_timeout(3000)

        # 滚动页面触发懒加载
        for i in range(4):
            page.evaluate(f"window.scrollTo(0, {i * 600})")
            page.wait_for_timeout(600)

        # ── 诊断：打印当前页面 URL ──
        print(f"\n[DIAGNOSTIC] 当前页面 URL:")
        print(f"  {page.url[:80]}")
        if "login" in page.url or "passport" in page.url:
            print("\n  ⚠️  检测到登录页，请先在浏览器登录京东！")
        elif "verify" in page.url or "captcha" in page.url.lower():
            print("\n  ⚠️  检测到验证码/安全验证页，请手动完成验证！")
        else:
            # 诊断截图
            diag_path = os.path.join(save_dir, "diagnostic.png")
            page.screenshot(path=diag_path, full_page=False)
            print(f"\n[DIAGNOSTIC] 诊断截图已保存")
            try:
                title = page.title()
                print(f"  页面标题: {title}")
            except:
                pass

        # 打印当前 URL，看是否跳转了
        print(f"[PAGE] 当前页面: {page.url[:70]}")
        if "login" in page.url or "passport" in page.url:
            print("\n" + "!"*55)
            print("  ⚠️  需要登录！请在浏览器中完成京东登录")
            print("  ⚠️  登录后页面会自动跳转，脚本会继续...")
            print("!"*55)
            try:
                page.wait_for_url("**item.jd.com**", timeout=120000)
                print("[OK] 登录成功，正在提取数据...")
            except:
                print("[WARN] 等待超时，请检查浏览器页面")

        # 提取数据
        print("\n" + "="*50)
        print("[EXTRACTING] 提取商品信息...")
        extractor = JDExtractor()
        data = extractor.extract(page, url)
        from datetime import datetime
        data["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 截图
        print("[SCREENSHOT] 截取页面...")
        shot_mgr = ScreenshotManager(page, save_dir)
        shot_mgr.screenshot_page()
        screenshot_path = os.path.join(save_dir, "screenshot.png")

        # 下载主图
        if data.get("main_images"):
            print(f"[IMAGES] 下载 {len(data['main_images'])} 张主图...")
            shot_mgr.download_main_images(data["main_images"])

        # OCR 兜底：店铺名未提取时触发
        if not data.get("shop_name") or data.get("shop_name") == "未提取":
            print("[OCR] 店铺名未提取，启用 OCR 识别...")
            ocr_shop = JDExtractor._ocr_shop_name(screenshot_path)
            if ocr_shop and ocr_shop != "未提取":
                data["shop_name"] = ocr_shop
                print(f"[OCR] OCR 识别到店铺名: {ocr_shop}")
            else:
                print("[OCR] OCR 也未能识别到店铺名")

        # 导出
        print("[EXCEL] 导出 Excel...")
        ExcelExporter(data, save_dir).export()

        print("[REPORT] 生成举报模板...")
        ReportGenerator(data, save_dir).generate()

        # 写入日志
        try:
            log_path = os.path.join(OUTPUT_DIR, "jd_log.txt")
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"[{data['timestamp']}] {data['platform']} | {data['title'][:40]} | {url}\n")
        except:
            pass

        # ── 输出结果 ──
        print(f"""
╔══════════════════════════════════════════════════════════╗
║                    ✅ 提取完成                            ║
╚══════════════════════════════════════════════════════════╝

[提取结果]
  平台：{data['platform']}
  标题：{data['title'][:50] if len(data['title']) > 50 else data['title']}
  价格：{data['price']}
  店铺：{data['shop_name']}
  卖家ID：{data['seller_id']}
  销量：{data['sales']}
  侵权检查：{data['infringement_check']}

[输出目录]
  {save_dir}

[文件列表]
  📊 jd_extraction.xlsx  - Excel 数据表
  📄 jd_report.txt        - 举报模板
  🖼️  screenshot.png      - 页面截图
  📁 main_images/         - 商品主图
""")
        browser.close()


if __name__ == "__main__":
    main()
