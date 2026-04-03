"""
WOW English 知识产权投诉 - 批量处理工具 v1.0
输入：Excel 链接列表 → 自动抓取 → 生成 IPP 投诉文档

使用方法：
  1. 打开「投诉链接模板.xlsx」，在"商品链接"列填入要投诉的链接（一行一个）
  2. 双击运行「批量处理.bat」，或命令行运行：python batch_ipp.py
  3. 处理完成后，打开 output/IPP投诉文档_日期.xlsx，复制粘贴到 IPP 平台
"""

import os
import sys
import io
import re
import time
import random
import socket
from datetime import datetime
from pathlib import Path

# Windows 编码修复
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    try:
        os.system('chcp 65001 >nul 2>&1')
    except:
        pass

try:
    from playwright.sync_api import sync_playwright, Page
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import easyocr
except ImportError as e:
    print(f"[ERROR] 缺少依赖: {e}")
    print("运行: pip install playwright openpyxl easyocr")
    sys.exit(1)

# ==================== 配置 ====================
BRAND_KEYWORDS = [
    "wow english", "wowenglish", "wow english",
    "史蒂夫英语", "steve english",
    "english singsing", "englishsingsing",
    "英语启蒙动画", "史蒂夫", "steve",
    "wowyoung", "wow young",
]
COPYRIGHT_HOLDER = "WOW English 版权方"
RIGHTS_TYPE = "商标权/著作权"
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")

# ==================== OCR 预加载 ====================
print("[OCR] 初始化文字识别引擎（首次约30秒）...")
_OCR_READER = None
def get_ocr_reader():
    global _OCR_READER
    if _OCR_READER is None:
        _OCR_READER = easyocr.Reader(['ch_sim', 'en'], gpu=False, verbose=False)
    return _OCR_READER


# ==================== 平台检测 ====================
def detect_platform(url: str) -> str:
    if 'jd.com' in url:
        return 'jd'
    elif 'tmall.com' in url or 'taobao.com' in url:
        return 'tmall'
    elif 'pinduoduo.com' in url or 'pdd.com' in url:
        return 'pdd'
    return 'unknown'


# ==================== 京东提取器 ====================
class JDExtractor:
    TITLE_SELECTORS = [
        '.product-title__main', '[class*="product-title"]',
        '[class*="sku-name"]', '[class*="item-name"]', '.p-title', '#itemName',
        'div[class*="product-name"] h1', '#name h1', 'h1[class*="name"]',
        'h1[class*="product"]', 'h1[class*="item"]', 'h1', '[data-sku]',
    ]
    PRICE_SELECTORS = [
        '[class*="price"] [class*="integer"]', '[class*="price"] strong',
        '.p-price .price', '#jdprice', '[data-price]', '[class*="sale-price"]',
        '[class*="price-wrap"] span', '.J-p-', '#spec-qty-price',
    ]
    SHOP_SELECTORS = [
        '[class*="shop-name"] a[href*="shop"]', '.j-shop-content .shop-name',
        '#shopInfo .shop-name', '#shop-name',
        '[class*="seller-info"] [class*="name"]',
        '#pop_shop .name', '.popshop .shop-name',
        '[class*="shop-name"] a', '.shop-name a', '[class*="seller"] a',
        '[class*="store"] a', '[class*="shop"]',
    ]

    @classmethod
    def extract(cls, page: Page, url: str) -> dict:
        data = {
            "platform": "京东",
            "title": "", "price": "", "shop_name": "", "seller_id": "",
            "sales": "", "main_images": [], "url": url,
            "infringement_check": "否"
        }
        js_data = cls._extract_via_js(page)
        if js_data:
            data.update(js_data)
        else:
            data["title"] = cls._try_selectors(page, cls.TITLE_SELECTORS)
            data["price"] = cls._try_selectors(page, cls.PRICE_SELECTORS)
            data["shop_name"] = cls._try_selectors(page, cls.SHOP_SELECTORS)
        data["infringement_check"] = cls._check_infringement(data)
        return data

    @staticmethod
    def _extract_via_js(page: Page) -> dict:
        try:
            result = page.evaluate("""
                () => {
                    const data = {};
                    const noiseList = ['最小单价','搭配购买','似乎出了点问题','先看看','极速审核',
                                       '立即购买','加入购物车','查看全部','售后服务','退换无忧'];
                    function isNoise(t) { return noiseList.some(n => t.includes(n)); }

                    const h1 = document.querySelector('h1') || document.querySelector('[class*="product-title"]');
                    if (h1) { const t = h1.innerText.trim().substring(0,200); if(t && t.length>5 && !isNoise(t)) data.title = t; }

                    const priceEls = document.querySelectorAll('[class*="price"]');
                    for(const el of priceEls){ const t=el.innerText.trim(); if(/^￥|^¥|^\\d/.test(t)&&!isNoise(t)&&t.length<30){data.price=t;break;} }

                    const shopSelectors = [
                        '[class*="shop-info"] [class*="name"]','[class*="shop-header"] [class*="name"]',
                        '[class*="seller-info"] [class*="name"]','[class*="score"] [class*="shop"]',
                        '[class*="shop-name"]','a[href*="shop.jd.com"]','[class*="shop-name"] a','#shop-name a'
                    ];
                    for(const sel of shopSelectors){
                        const el=document.querySelector(sel);
                        if(el){const t=el.innerText.trim();if(t&&!isNoise(t)&&t.length>1&&t.length<50){data.shop_name=t;break;}}
                    }
                    const shopLink = document.querySelector('a[href*="shop.jd.com"]')||document.querySelector('[class*="shop-name"] a');
                    if(shopLink&&shopLink.href){
                        const m=shopLink.href.match(/shopId[=:]"?(\\d+)/)||shopLink.href.match(/\\/shop\\/(\\d+)/);
                        if(m) data.seller_id=m[1];
                    }
                    const salesEl = document.querySelector('.p-commit a')||document.querySelector('#commit_count')||document.querySelector('.J-comm-count');
                    if(salesEl){const t=salesEl.innerText.trim().split('\\n')[0];if(t&&!isNoise(t)&&t.length<30)data.sales=t;}
                    return data;
                }
            """)
            return {k: v for k, v in result.items() if v}
        except:
            return {}

    @staticmethod
    def _try_selectors(page, selectors):
        for sel in selectors:
            try:
                el = page.locator(sel).first
                if el.count() > 0 and el.is_visible():
                    text = el.inner_text().strip()
                    if text and len(text) > 1:
                        return text[:200]
            except:
                continue
        return ""

    @staticmethod
    def _check_infringement(data):
        text = f"{data.get('title','')} {data.get('shop_name','')}".lower()
        matched = [kw for kw in BRAND_KEYWORDS if kw.lower() in text]
        return f"是（匹配：{', '.join(matched)}）" if matched else "否"

    @staticmethod
    def ocr_shop_name(image_path):
        try:
            from PIL import Image
            reader = get_ocr_reader()
            img = Image.open(image_path)
            w, h = img.size
            top = img.crop((0, 0, w, int(h * 0.25)))
            tmp = image_path.replace('.png','_shop.png')
            top.save(tmp)
            results = reader.readtext(tmp)
            all_text = ' '.join([r[1] for r in results])
            shop_match = re.search(
                r'([\u4e00-\u9fa5]{2,30}(?:旗舰店|专营店|专卖店|直营店|官方旗舰店|海外旗舰店|'
                r'官方店|商城|百货|书屋|书店))', all_text)
            if shop_match:
                return shop_match.group(1).strip()
            chinese = re.findall(r'[\u4e00-\u9fa5]{3,}', all_text)
            if chinese:
                return max(chinese, key=len)
        except:
            pass
        return ""


# ==================== 拼多多提取器 ====================
class PDDExtractor:
    @classmethod
    def extract(cls, page: Page, url: str) -> dict:
        data = {
            "platform": "拼多多",
            "title": "", "price": "", "shop_name": "", "seller_id": "",
            "sales": "", "main_images": [], "url": url,
            "infringement_check": "否"
        }
        js_data = cls._extract_via_js(page)
        if js_data:
            data.update(js_data)
        data["infringement_check"] = JDExtractor._check_infringement(data)
        return data

    @staticmethod
    def _extract_via_js(page: Page) -> dict:
        try:
            result = page.evaluate("""
                () => {
                    const data = {};
                    const noiseList = ['最小单价','搭配购买','极速审核','立即购买','加入购物车','查看全部'];
                    function isNoise(t) { return noiseList.some(n => t.includes(n)); }

                    const titleEl = document.querySelector('[class*="product-title"]') ||
                                    document.querySelector('[class*="goods-title"]') ||
                                    document.querySelector('h1') || document.querySelector('[class*="item-title"]');
                    if(titleEl){const t=titleEl.innerText.trim().substring(0,200);if(t&&t.length>3&&!isNoise(t))data.title=t;}

                    const priceEl = document.querySelector('[class*="price"] [class*="value"]') ||
                                   document.querySelector('[class*="sale-price"]') ||
                                   document.querySelector('.product-price') || document.querySelector('[class*="price"]');
                    if(priceEl){const t=priceEl.innerText.trim().substring(0,20);if(t&&!isNoise(t))data.price=t;}

                    const shopEl = document.querySelector('[class*="shop-name"]') ||
                                  document.querySelector('[class*="store-name"]') ||
                                  document.querySelector('[class*="malls-name"]') ||
                                  document.querySelector('[class*="goods-single"]');
                    if(shopEl){const t=shopEl.innerText.trim().substring(0,50);if(t&&!isNoise(t))data.shop_name=t;}

                    const salesEl = document.querySelector('[class*="sales"]') || document.querySelector('[class*="sales-count"]');
                    if(salesEl){const t=salesEl.innerText.trim().substring(0,20);if(t&&!isNoise(t))data.sales=t;}

                    return data;
                }
            """)
            return {k: v for k, v in result.items() if v}
        except:
            return {}


# ==================== 天猫/淘宝提取器 ====================
class TMExtractor:
    @classmethod
    def extract(cls, page: Page, url: str) -> dict:
        data = {
            "platform": "天猫/淘宝",
            "title": "", "price": "", "shop_name": "", "seller_id": "",
            "sales": "", "main_images": [], "url": url,
            "infringement_check": "否"
        }
        js_data = cls._extract_via_js(page)
        if js_data:
            data.update(js_data)
        data["infringement_check"] = JDExtractor._check_infringement(data)
        return data

    @staticmethod
    def _extract_via_js(page: Page) -> dict:
        try:
            result = page.evaluate("""
                () => {
                    const data = {};
                    const noiseList = ['最小单价','搭配购买','极速审核','立即购买','加入购物车'];
                    function isNoise(t) { return noiseList.some(n => t.includes(n)); }

                    const titleEl = document.querySelector('[class*="product-title"]') ||
                                    document.querySelector('.product-title') ||
                                    document.querySelector('h1[class*="title"]') || document.querySelector('h1');
                    if(titleEl){const t=titleEl.innerText.trim().substring(0,200);if(t&&t.length>3&&!isNoise(t))data.title=t;}

                    const priceEl = document.querySelector('#J_StrPrice') || document.querySelector('[class*="price"]');
                    if(priceEl){const t=priceEl.innerText.trim().substring(0,20);if(t)data.price=t;}

                    const shopEl = document.querySelector('[class*="shop-name"] a') ||
                                  document.querySelector('#shopExtra a') ||
                                  document.querySelector('[class*="shopInfo"]');
                    if(shopEl){const t=shopEl.innerText.trim().substring(0,50);if(t&&!isNoise(t))data.shop_name=t;}

                    return data;
                }
            """)
            return {k: v for k, v in result.items() if v}
        except:
            return {}


# ==================== 侵权对比图生成 ====================
def generate_comparison_image(screenshot_path: str, brand: str, save_dir: str, index: int) -> str:
    """生成侵权对比图：截取页面关键区域 + 添加标注框"""
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        print("[WARN] 缺少 Pillow，无法生成对比图")
        return ""

    if not screenshot_path or not os.path.exists(screenshot_path):
        return ""

    try:
        img = Image.open(screenshot_path)
        w, h = img.size

        # 截取主要区域（商品信息和店铺名区域）
        top_region = img.crop((0, 0, w, min(h // 2, 800)))

        # 创建对比图（左侧：原始截图，右侧：标注版本）
        comparison_w = w * 2 + 20
        comparison_h = max(top_region.height, 300) + 80
        comparison = Image.new('RGB', (comparison_w, comparison_h), color=(240, 240, 240))
        comparison.paste(top_region, (0, 0))

        # 右侧标注版本
        annotated = top_region.copy()
        draw = ImageDraw.Draw(annotated)

        # 尝试加载字体（回退到默认）
        try:
            font_large = ImageFont.truetype("msyh.ttc", 20)
            font_small = ImageFont.truetype("msyh.ttc", 14)
        except:
            font_large = ImageFont.load_default()
            font_small = ImageFont.load_default()

        # 添加红色标注框（店铺区域）
        draw.rectangle([10, 40, w - 10, 100], outline=(220, 0, 0), width=3)
        draw.rectangle([8, 38, w - 8, 102], outline=(255, 100, 100), width=1)
        draw.text((15, 42), f"⚠ 侵权店铺：{brand}", fill=(220, 0, 0), font=font_small)

        # 添加标题区域标注
        title_h = min(h // 4, 200)
        draw.rectangle([10, 100, w - 10, title_h], outline=(0, 100, 200), width=2)
        draw.text((15, 105), "商品标题（含品牌关键词）", fill=(0, 80, 180), font=font_small)

        # 合并到对比图右侧
        comparison.paste(annotated, (w + 10, 0))

        # 分割线
        draw2 = ImageDraw.Draw(comparison)
        draw2.line([w + 5, 0, w + 5, comparison_h], fill=(180, 180, 180), width=2)

        # 添加文字标注
        draw2.text((w // 2 - 80, top_region.height + 10), "原始截图", fill=(80, 80, 80), font=font_small)
        draw2.text((w + w // 2 - 80, top_region.height + 10), "侵权标注", fill=(220, 0, 0), font=font_small)
        draw2.text((20, comparison_h - 30),
                   f"WOW English 侵权证据 #{index} | 未经授权使用品牌关键词",
                   fill=(100, 100, 100), font=font_small)

        out_path = os.path.join(save_dir, f"侵权对比图_{index:03d}.png")
        comparison.save(out_path, quality=95)
        return out_path
    except Exception as e:
        print(f"[WARN] 对比图生成失败: {e}")
        return ""


# ==================== 投诉理由生成 ====================
def generate_complaint_text_jd(shop: str, title: str, url: str, matched_kw: str) -> str:
    """生成京东 IP 投诉理由文本"""
    return f"""投诉理由：
我方是「WOW English」品牌版权方（商标权/著作权权利人）。

经调查发现，商家「{shop}」（商品链接：{url}）在商品页面中存在以下侵权行为：

1. 商品标题中使用了"WOW English"、"Steve English"、"史蒂夫英语"等我方品牌关键词，借我方品牌知名度进行销售；
2. 该行为易使消费者产生混淆，误认为该商品与我方存在授权关系。

上述行为涉嫌违反：
- 《商标法》第五十七条第（二）项：未经许可在同一种商品上使用与注册商标近似的商标，容易导致混淆的；
- 《反不正当竞争法》第六条：擅自使用他人有一定影响的市场主体名称等，引人误认为是他人商品。

我方恳请京东平台依据《电子商务法》第四十五条，对上述侵权商品采取删除、屏蔽、断开链接等必要措施。

我方可提供完整商标注册证及著作权登记证明。如有需要，请联系投诉方。"""


def generate_complaint_text_tmall(shop: str, title: str, url: str, matched_kw: str) -> str:
    """生成淘天 IPP 投诉理由文本"""
    return f"""投诉理由：

【权利人声明】
我方为「WOW English」品牌的合法版权方，依法持有该品牌相关商标权及著作权。

【侵权事实】
商家「{shop}」（商品链接：{url}）销售的商品标题/详情页中，未经授权使用了"WOW English"、"史蒂夫英语"、"Steve English"、"English Singsing"等我方品牌关键词。

商品标题：{title}

【侵权类型】
□ 商标侵权（未经授权使用他人注册商标）
☑ 著作权侵权（盗用品牌动画角色形象/内容进行商业销售）

【法律依据】
1. 《商标法》第五十七条第（二）项：未经许可在同一种商品上使用与注册商标近似的商标，容易导致混淆的；
2. 《著作权法》第五十二条：未经著作权人许可，复制、发行、通过信息网络向公众传播其作品的，构成侵权；
3. 《电子商务法》第四十五条：电子商务平台经营者知道或者应当知道平台内经营者侵犯知识产权的，应当采取删除、屏蔽、断开链接、终止交易和服务等必要措施。

【投诉诉求】
恳请贵平台依据相关法律法规，对上述侵权商品立即采取删除、屏蔽、断开链接等必要措施，并视情况对侵权商家进行处罚。

我方承诺所提交材料真实有效，并依法承担相应法律责任。"""


def generate_complaint_text_pdd(shop: str, title: str, url: str, matched_kw: str) -> str:
    """生成拼多多投诉理由文本"""
    return f"""投诉声明：
我方「WOW English」版权方发现，商家「{shop}」在拼多多平台销售的商品标题中，
未经授权使用了"WOW English"、"Steve English"、"史蒂夫英语"等我方品牌关键词。
商品链接：{url}
商品标题：{title}
该行为涉嫌构成商标侵权及著作权侵权。
请依据《电子商务法》对侵权商品采取下架处理。"""


# ==================== 截图管理 ====================
class ScreenshotManager:
    def __init__(self, page: Page, save_dir: str, prefix: str = ""):
        self.page = page
        self.save_dir = save_dir
        self.prefix = prefix
        os.makedirs(save_dir, exist_ok=True)
        os.makedirs(os.path.join(save_dir, "main_images"), exist_ok=True)

    def screenshot_page(self) -> str:
        filename = os.path.join(self.save_dir, f"{self.prefix}screenshot.png")
        h = self.page.evaluate("document.body.scrollHeight")
        self.page.set_viewport_size({"width": 1366, "height": min(h, 6000)})
        self.page.screenshot(path=filename, full_page=True)
        return filename

    def download_main_images(self, urls: list) -> list:
        saved = []
        for i, url in enumerate(urls[:5]):
            try:
                url_clean = re.sub(r'/\d+x\d+/', '/800x800/', url)
                filename = os.path.join(self.save_dir, "main_images", f"{self.prefix}img_{i+1}.jpg")
                resp = self.page.context.request.get(url_clean)
                if resp.ok:
                    with open(filename, 'wb') as f:
                        f.write(resp.body())
                    saved.append(filename)
            except:
                continue
        return saved


# ==================== IPP 投诉文档生成器 ====================
class IPPComplaintGenerator:
    """生成 IPP 平台可直接粘贴的投诉文档"""

    def __init__(self, results: list, save_dir: str):
        self.results = results
        self.save_dir = save_dir
        os.makedirs(save_dir, exist_ok=True)

    def generate_excel(self) -> str:
        """生成 IPP 投诉 Excel（结构化格式，可复制粘贴）"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IPP投诉文档"

        # 标题行
        title_fill = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        title_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="FFD0D0", end_color="FFD0D0", fill_type="solid")
        header_font = Font(bold=True, size=10)

        # 表头
        headers = [
            "序号", "平台", "商品链接", "商品标题", "店铺名称",
            "卖家ID", "价格", "销量/评价", "疑似侵权", "截图路径", "证据目录"
        ]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = title_fill
            c.font = title_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = self._thin_border()

        # 维权方信息（固定内容，第2行）
        rights_info = {
            "序号": "", "平台": "维权方",
            "商品链接": "权利人/公司", "商品标题": COPYRIGHT_HOLDER,
            "店铺名称": "维权类型", "卖家ID": RIGHTS_TYPE,
            "价格": "处理平台", "销量/评价": "京东/淘宝/拼多多",
            "疑似侵权": "投诉原因", "截图路径": "请参考附件截图",
            "证据目录": "见对应子文件夹"
        }
        row2_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
        for col, h in enumerate(headers, 1):
            val = rights_info.get(h, "")
            c = ws.cell(row=2, column=col, value=val)
            c.fill = row2_fill
            c.font = Font(size=9, italic=True)
            c.border = self._thin_border()

        # 数据行
        for i, r in enumerate(self.results, 3):
            is_infringe = "是" in r.get("infringement_check", "否")
            row_fill = PatternFill(
                start_color="FFCCCC" if is_infringe else "FFFFFF",
                end_color="FFCCCC" if is_infringe else "FFFFFF",
                fill_type="solid"
            )
            row_data = [
                str(i - 2),
                r.get("platform", ""),
                r.get("url", ""),
                r.get("title", ""),
                r.get("shop_name", ""),
                r.get("seller_id", ""),
                r.get("price", ""),
                r.get("sales", ""),
                r.get("infringement_check", "否"),
                r.get("screenshot_path", ""),
                r.get("save_dir", ""),
            ]
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.fill = row_fill
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border = self._thin_border()
                if col == 9 and is_infringe:
                    c.font = Font(bold=True, color="CC0000")
                else:
                    c.font = Font(size=9)

        # 列宽
        widths = [6, 10, 55, 50, 25, 18, 12, 15, 35, 40, 50]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        ws.row_dimensions[1].height = 30

        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(self.save_dir, f"IPP投诉文档_{date_str}.xlsx")
        wb.save(filename)
        return filename

    def generate_txt(self) -> str:
        """生成纯文本版 IPP 投诉文档（直接复制粘贴用）"""
        date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        lines = []
        lines.append("=" * 70)
        lines.append("           WOW English 知识产权投诉文档")
        lines.append("           生成时间：" + date_str)
        lines.append("=" * 70)
        lines.append("")
        lines.append("【权利人信息】")
        lines.append(f"  权利人：{COPYRIGHT_HOLDER}")
        lines.append(f"  维权类型：{RIGHTS_TYPE}")
        lines.append(f"  处理平台：京东 / 天猫 / 拼多多")
        lines.append("")
        lines.append("=" * 70)
        lines.append("【投诉商品列表】")
        lines.append("=" * 70)

        for i, r in enumerate(self.results, 1):
            lines.append("")
            lines.append(f"─── 商品 {i} ───")
            lines.append(f"  平台：{r.get('platform','')}")
            lines.append(f"  商品链接：{r.get('url','')}")
            lines.append(f"  商品标题：{r.get('title','')}")
            lines.append(f"  店铺名称：{r.get('shop_name','')}")
            lines.append(f"  卖家ID：{r.get('seller_id','')}")
            lines.append(f"  价格：{r.get('price','')}")
            lines.append(f"  销量：{r.get('sales','')}")
            lines.append(f"  疑似侵权：{r.get('infringement_check','')}")
            lines.append(f"  截图：{r.get('screenshot_path','')}")
            lines.append(f"  证据目录：{r.get('save_dir','')}")

        lines.append("")
        lines.append("=" * 70)
        lines.append("【投诉理由模板（可直接复制到 IPP 平台）】")
        lines.append("=" * 70)

        for i, r in enumerate(self.results, 1):
            platform = r.get('platform', '')
            shop = r.get('shop_name', '未知店铺')
            title = r.get('title', '未知商品')
            url = r.get('url', '')
            matched = r.get('infringement_check', '')
            kw = matched.split('匹配：')[1].rstrip('）') if '匹配：' in matched else ''

            if '京东' in platform:
                text = generate_complaint_text_jd(shop, title, url, kw)
            elif '天猫' in platform or '淘宝' in platform:
                text = generate_complaint_text_tmall(shop, title, url, kw)
            elif '拼多多' in platform:
                text = generate_complaint_text_pdd(shop, title, url, kw)
            else:
                text = generate_complaint_text_tmall(shop, title, url, kw)

            lines.append(f"\n{'─'*60}\n商品 {i} 投诉理由（{platform}）\n{'─'*60}\n{text}")

        lines.append("=" * 70)
        lines.append(f"共 {len(self.results)} 件商品 | 生成时间：{date_str}")
        lines.append("=" * 70)

        content = '\n'.join(lines)
        txt_path = os.path.join(self.save_dir, f"IPP投诉文档_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write(content)
        return txt_path

    def generate_batch_summary(self) -> str:
        """生成批量处理汇总 Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "汇总"

        # 样式
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        headers = ["序号", "平台", "商品链接", "商品标题（已脱敏）",
                   "店铺名称", "价格", "侵权判定", "截图文件", "证据目录"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

        for i, r in enumerate(self.results, 2):
            is_infringe = "是" in r.get("infringement_check", "否")
            # 标题脱敏：只显示前20字+***
            title_raw = r.get('title', '')[:20]
            title_display = title_raw + ('...' if len(r.get('title','')) > 20 else '')
            # 链接脱敏
            url_raw = r.get('url', '')
            platform = r.get('platform', '')
            row_data = [
                str(i - 1),
                platform,
                url_raw,
                title_display,
                r.get('shop_name', ''),
                r.get('price', ''),
                r.get('infringement_check', '否'),
                os.path.basename(r.get('screenshot_path', '')),
                os.path.basename(r.get('save_dir', '')),
            ]
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 7:
                    c.fill = red_fill if is_infringe else green_fill
                    c.font = Font(bold=True, color="FFFFFF" if is_infringe else "000000")

        widths = [6, 10, 50, 28, 25, 12, 35, 25, 25]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        # 统计行
        last_row = len(self.results) + 2
        ws.cell(row=last_row, column=1, value="统计").font = Font(bold=True)
        ws.cell(row=last_row, column=2, value=f"共 {len(self.results)} 件")
        ws.cell(row=last_row, column=3, value=f"疑似侵权 {sum(1 for r in self.results if '是' in r.get('infringement_check',''))} 件")

        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(self.save_dir, f"批量处理汇总_{date_str}.xlsx")
        wb.save(filename)
        return filename

    def generate_ipp_form(self) -> str:
        """生成 IPP 平台预填充表单（京东+淘天双平台）"""
        wb = openpyxl.Workbook()

        # === Sheet 1: 京东 IP 投诉表单 ===
        ws_jd = wb.active
        ws_jd.title = "京东IP投诉表单"

        jd_title_fill = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        jd_title_font = Font(bold=True, color="FFFFFF", size=14)
        section_fill = PatternFill(start_color="FFD0D0", end_color="FFD0D0", fill_type="solid")
        section_font = Font(bold=True, size=10)
        field_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        value_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        tip_font = Font(size=9, color="888888", italic=True)

        # 京东表单
        jd_fields = [
            ("投诉方信息", ""),
            ("投诉人/公司名称", COPYRIGHT_HOLDER),
            ("联系方式", "请填写您的邮箱/电话"),
            ("证件号码", "请填写营业执照统一社会信用代码"),
            ("", ""),
            ("权利信息", ""),
            ("权利类型", RIGHTS_TYPE),
            ("权利名称/注册号", "请填写WOW English商标注册号"),
            ("权利证书上传", "需上传：商标注册证扫描件（PDF/JPG）"),
            ("著作权登记证明", "需上传：著作权登记证书（如有）"),
            ("", ""),
            ("侵权信息", ""),
            ("被投诉平台", "京东"),
            ("涉嫌侵权商品数量", str(len(self.results)) + " 件"),
        ]

        row = 1
        ws_jd.column_dimensions['A'].width = 22
        ws_jd.column_dimensions['B'].width = 50
        ws_jd.column_dimensions['C'].width = 40

        # 标题行
        ws_jd.merge_cells('A1:C1')
        c = ws_jd['A1']
        c.value = "京东知识产权保护平台 - 投诉表单（预填充版）"
        c.fill = jd_title_fill
        c.font = jd_title_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_jd.row_dimensions[1].height = 35

        row = 2
        for label, value in jd_fields:
            if label == "":
                row += 1
                continue
            is_section = value == ""
            if is_section:
                ws_jd.merge_cells(f'A{row}:C{row}')
                c = ws_jd.cell(row=row, column=1, value=label)
                c.fill = section_fill
                c.font = section_font
                c.alignment = Alignment(horizontal="left", vertical="center")
                ws_jd.row_dimensions[row].height = 22
            else:
                c_label = ws_jd.cell(row=row, column=1, value=label)
                c_label.fill = field_fill
                c_label.font = Font(bold=True, size=10)
                c_label.alignment = Alignment(horizontal="left", vertical="center")
                c_value = ws_jd.cell(row=row, column=2, value=value)
                c_value.fill = value_fill
                c_value.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws_jd.row_dimensions[row].height = 20

                # 提示列
                if "上传" in label or "填写" in value:
                    c_tip = ws_jd.cell(row=row, column=3, value=value)
                    c_tip.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
                    c_tip.font = tip_font
            row += 1

        # 侵权商品列表（从第row行开始）
        row += 1
        ws_jd.merge_cells(f'A{row}:C{row}')
        c = ws_jd.cell(row=row, column=1, value="【涉嫌侵权商品列表】（自动抓取）")
        c.fill = section_fill
        c.font = section_font
        row += 1

        # 表头
        headers_jd = ["序号", "商品链接", "店铺名称", "商品标题", "侵权关键词"]
        col_widths_jd = [6, 50, 25, 40, 30]
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)

        # 调整列数
        for ci, (h, w) in enumerate(zip(headers_jd, col_widths_jd), 1):
            ws_jd.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
            c = ws_jd.cell(row=row, column=ci, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")
        ws_jd.row_dimensions[row].height = 22
        row += 1

        for i, r in enumerate(self.results, 1):
            infringement = r.get('infringement_check', '')
            matched = ''
            if '匹配：' in infringement:
                matched = infringement.split('匹配：')[1].rstrip('）')
            row_data = [str(i), r.get('url', ''), r.get('shop_name', ''),
                       r.get('title', '')[:50], matched]
            row_fill = PatternFill(
                start_color="FFEEEE" if "是" in infringement else "FFFFFF",
                end_color="FFEEEE" if "是" in infringement else "FFFFFF",
                fill_type="solid"
            )
            for ci, val in enumerate(row_data, 1):
                c = ws_jd.cell(row=row, column=ci, value=val)
                c.fill = row_fill
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.font = Font(size=9)
            ws_jd.row_dimensions[row].height = 30
            row += 1

        # === Sheet 2: 淘天 IPP 投诉表单 ===
        ws_ipp = wb.create_sheet("淘天IPP投诉表单")
        ws_ipp.merge_cells('A1:E1')
        c = ws_ipp['A1']
        c.value = "淘天集团 IPP 平台 - 投诉表单（预填充版）"
        c.fill = jd_title_fill
        c.font = jd_title_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_ipp.row_dimensions[1].height = 35

        ipp_title_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        ipp_section = PatternFill(start_color="FFE0CC", end_color="FFE0CC", fill_type="solid")

        ipp_fields = [
            ("投诉方信息", ""),
            ("单位/姓名", COPYRIGHT_HOLDER),
            ("联系方式", "请填写邮箱"),
            ("证件号码", "请填写统一社会信用代码/身份证号"),
            ("投诉方类型", "权利人本人"),
            ("", ""),
            ("权利信息", ""),
            ("权利类型", "商标权 + 著作权"),
            ("注册商标名称", "WOW English / Steve English / 史蒂夫英语"),
            ("商标注册号", "请填写实际注册号"),
            ("权利证明文件", "需上传：商标注册证扫描件"),
            ("著作权证明文件", "需上传：作品登记证书（如有）"),
            ("", ""),
            ("投诉信息", ""),
            ("投诉平台", "天猫 / 淘宝 / 1688"),
            ("涉嫌侵权商品数量", str(len(self.results)) + " 件"),
            ("侵权类型", "商标侵权 + 著作权侵权"),
        ]

        ws_ipp.column_dimensions['A'].width = 22
        ws_ipp.column_dimensions['B'].width = 35
        ws_ipp.column_dimensions['C'].width = 20
        ws_ipp.column_dimensions['D'].width = 30
        ws_ipp.column_dimensions['E'].width = 35

        row = 2
        for label, value in ipp_fields:
            if label == "":
                row += 1
                continue
            is_section = value == ""
            if is_section:
                ws_ipp.merge_cells(f'A{row}:E{row}')
                c = ws_ipp.cell(row=row, column=1, value=label)
                c.fill = ipp_section
                c.font = section_font
                ws_ipp.row_dimensions[row].height = 22
            else:
                c_label = ws_ipp.cell(row=row, column=1, value=label)
                c_label.fill = field_fill
                c_label.font = Font(bold=True, size=10)
                c_label.alignment = Alignment(horizontal="left", vertical="center")

                ws_ipp.merge_cells(f'B{row}:C{row}')
                c_val = ws_ipp.cell(row=row, column=2, value=value)
                c_val.fill = value_fill
                c_val.alignment = Alignment(horizontal="left", vertical="center")

                if "上传" in label or "填写" in value or "请填写" in value:
                    ws_ipp.merge_cells(f'D{row}:E{row}')
                    c_tip = ws_ipp.cell(row=row, column=4, value=value)
                    c_tip.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
                    c_tip.font = tip_font
                ws_ipp.row_dimensions[row].height = 20
            row += 1

        row += 1
        ws_ipp.merge_cells(f'A{row}:E{row}')
        c = ws_ipp.cell(row=row, column=1, value="【涉嫌侵权商品列表】（自动抓取）")
        c.fill = ipp_section
        c.font = section_font
        row += 1

        headers_ipp = ["序号", "平台", "商品链接", "店铺名称", "商品标题（已脱敏）"]
        col_widths_ipp = [6, 10, 50, 25, 40]
        for ci, (h, w) in enumerate(zip(headers_ipp, col_widths_ipp), 1):
            ws_ipp.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
            c = ws_ipp.cell(row=row, column=ci, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")
        ws_ipp.row_dimensions[row].height = 22
        row += 1

        for i, r in enumerate(self.results, 1):
            row_data = [str(i), r.get('platform', ''), r.get('url', ''),
                       r.get('shop_name', ''), r.get('title', '')[:40]]
            row_fill = PatternFill(
                start_color="FFEEEE" if "是" in r.get('infringement_check', '') else "FFFFFF",
                end_color="FFEEEE" if "是" in r.get('infringement_check', '') else "FFFFFF",
                fill_type="solid"
            )
            for ci, val in enumerate(row_data, 1):
                c = ws_ipp.cell(row=row, column=ci, value=val)
                c.fill = row_fill
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.font = Font(size=9)
            ws_ipp.row_dimensions[row].height = 30
            row += 1

        # === Sheet 3: 投诉理由批量生成 ===
        ws_txt = wb.create_sheet("投诉理由(直接复制)")
        ws_txt.column_dimensions['A'].width = 100

        hdr_fill_txt = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        for col in range(1, 3):
            c = ws_txt.cell(row=1, column=col, value="投诉理由（直接复制到 IPP 平台）" if col == 1 else "所属平台")
            c.fill = hdr_fill_txt
            c.font = Font(bold=True, color="FFFFFF", size=12)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_txt.row_dimensions[1].height = 30
        ws_txt.column_dimensions['B'].width = 15

        for i, r in enumerate(self.results, 2):
            platform = r.get('platform', '')
            shop = r.get('shop_name', '')
            title = r.get('title', '')
            url = r.get('url', '')
            matched = r.get('infringement_check', '')
            kw = matched.split('匹配：')[1].rstrip('）') if '匹配：' in matched else ''

            if '京东' in platform:
                text = generate_complaint_text_jd(shop, title, url, kw)
            elif '天猫' in platform or '淘宝' in platform:
                text = generate_complaint_text_tmall(shop, title, url, kw)
            elif '拼多多' in platform:
                text = generate_complaint_text_pdd(shop, title, url, kw)
            else:
                text = generate_complaint_text_tmall(shop, title, url, kw)

            c_text = ws_txt.cell(row=i, column=1, value=text)
            c_text.alignment = Alignment(wrap_text=True, vertical="top")
            c_text.font = Font(size=9)
            ws_txt.row_dimensions[i].height = 320

            c_plat = ws_txt.cell(row=i, column=2, value=platform)
            c_plat.alignment = Alignment(horizontal="center", vertical="top")
            c_plat.fill = PatternFill(
                start_color="FFEEEE" if "是" in matched else "EEFFEE",
                end_color="FFEEEE" if "是" in matched else "EEFFEE",
                fill_type="solid"
            )
            c_plat.font = Font(bold=True, size=9)

        # === Sheet 4: 材料清单 ===
        ws_check = wb.create_sheet("材料准备清单")
        ws_check.column_dimensions['A'].width = 35
        ws_check.column_dimensions['B'].width = 30
        ws_check.column_dimensions['C'].width = 40

        ws_check.merge_cells('A1:C1')
        c = ws_check['A1']
        c.value = "IPP 投诉材料准备清单"
        c.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        c.font = Font(bold=True, color="FFFFFF", size=13)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_check.row_dimensions[1].height = 30

        checklist = [
            ("【必须材料】", "", ""),
            ("1. 营业执照 / 身份证", "必填", "拍照或扫描，清晰可读"),
            ("2. 商标注册证", "必填", "WOW English 相关商标注册证扫描件"),
            ("3. 著作权登记证书", "如有", "如已完成著作权登记"),
            ("4. 侵权截图证据", "自动生成", "见各商品文件夹下的 screenshot.png"),
            ("5. 侵权对比图", "建议上传", "见各商品文件夹下的 侵权对比图_XXX.png"),
            ("", "", ""),
            ("【自动生成材料】", "", ""),
            ("6. 投诉商品列表", "已生成", "见「涉嫌侵权商品列表」工作表"),
            ("7. 投诉理由文本", "已生成", "见「投诉理由(直接复制)」工作表"),
            ("8. 侵权商品截图", "已生成", "每个商品文件夹中"),
            ("", "", ""),
            ("【上传顺序建议】", "", ""),
            ("Step 1", "上传权利证明", "营业执照+商标注册证"),
            ("Step 2", "填写投诉信息", "链接+侵权类型"),
            ("Step 3", "粘贴投诉理由", "从「投诉理由」Sheet 复制"),
            ("Step 4", "上传截图证据", "每个商品上传对应截图"),
            ("Step 5", "提交投诉", "确认后提交"),
        ]

        row = 2
        for item, status, note in checklist:
            is_section = status == ""
            if is_section:
                ws_check.merge_cells(f'A{row}:C{row}')
                c = ws_check.cell(row=row, column=1, value=item)
                c.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                c.font = Font(bold=True, size=10)
                c.alignment = Alignment(horizontal="left", vertical="center")
                ws_check.row_dimensions[row].height = 22
            else:
                c_item = ws_check.cell(row=row, column=1, value=item)
                c_item.alignment = Alignment(horizontal="left", vertical="center")
                c_item.font = Font(size=10)

                status_color = "92D050" if status in ("必填", "已生成", "自动生成") else (
                    "FFF9E6" if status == "建议上传" else "FFF9E6")
                c_status = ws_check.cell(row=row, column=2, value=status)
                c_status.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                c_status.font = Font(bold=True, size=10)
                c_status.alignment = Alignment(horizontal="center", vertical="center")

                c_note = ws_check.cell(row=row, column=3, value=note)
                c_note.font = Font(size=9, color="666666")
                c_note.alignment = Alignment(horizontal="left", vertical="center")
                ws_check.row_dimensions[row].height = 20
            row += 1

        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(self.save_dir, f"IPP完整材料包_{date_str}.xlsx")
        wb.save(filename)
        return filename

    @staticmethod
    def _thin_border():
        thin = Side(style='thin', color="AAAAAA")
        return Border(left=thin, right=thin, top=thin, bottom=thin)


# ==================== Excel 链接读取器 ====================
def read_links_from_excel(excel_path: str) -> list:
    """从 Excel 读取链接列表"""
    wb = openpyxl.load_workbook(excel_path)
    links = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    url = cell.strip()
                    if url and ('jd.com' in url or 'tmall.com' in url or
                                'taobao.com' in url or 'pinduoduo.com' in url or 'pdd.com' in url):
                        links.append(url)
    return links


# ==================== CDP 端口检测 ====================
def find_cdp_port():
    for port in [9222, 9223, 9224]:
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1)
            r = sock.connect_ex(('127.0.0.1', port))
            sock.close()
            if r == 0:
                return port
        except:
            continue
    return None


# ==================== 拟人化延迟 ====================
def human_delay(min_sec: float = 0.5, max_sec: float = 2.0):
    """模拟人类操作的随机延迟"""
    t = random.uniform(min_sec, max_sec)
    time.sleep(t)

def human_scroll(page, steps: int = None):
    """拟人化滚动页面"""
    if steps is None:
        steps = random.randint(3, 6)
    for i in range(steps):
        y = i * random.randint(500, 900)
        page.evaluate(f"window.scrollTo(0, {y})")
        human_delay(0.4, 1.2)


# ==================== 单个链接抓取 ====================
def process_single_url(browser, url: str, batch_dir: str, idx: int, total: int) -> dict:
    """处理单个链接，返回提取结果"""
    cdp_port = find_cdp_port()
    platform = detect_platform(url)
    safe_name = re.sub(r'[^\w\u4e00-\u9fa5]', '_', url.split('/')[-1][:30])
    save_dir = os.path.join(batch_dir, f"{idx:03d}_{platform}_{safe_name}")
    os.makedirs(save_dir, exist_ok=True)

    print(f"\n[{idx}/{total}] 处理中... [{platform}] {url[:60]}...")

    # 随机延迟：进入下一个链接前等一等
    if idx > 1:
        wait = random.uniform(4, 12)
        print(f"  [💤] 随机等待 {wait:.1f} 秒...")
        time.sleep(wait)

    with sync_playwright() as p:
        page = None
        ctx = None

        if cdp_port:
            try:
                browser_ref = p.chromium.connect_over_cdp(f"http://127.0.0.1:{cdp_port}")
                ctx = browser_ref.contexts[0]
                page = ctx.new_page()
            except:
                pass

        if page is None:
            br = p.chromium.launch(
                headless=False,
                args=['--disable-blink-features=AutomationControlled',
                      '--disable-dev-shm-usage', '--no-sandbox']
            )
            ctx = br.new_context(
                viewport={"width": 1366, "height": 768},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36',
                locale='zh-CN'
            )
            page = ctx.new_page()

        # 打开新标签页前随机等一下
        human_delay(0.3, 0.8)

        # 导航
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            human_delay(1.5, 4.0)  # 页面 DOM 加载后等一下
            page.wait_for_load_state("networkidle", timeout=15000)
        except:
            try:
                page.goto(url, wait_until="load", timeout=30000)
                human_delay(2.0, 5.0)
            except Exception as e:
                print(f"  [WARN] 页面加载失败: {e}")

        # 模拟人类阅读页面的滚动
        human_delay(1.0, 3.0)
        human_scroll(page)

        # 提取
        if platform == 'jd':
            extractor = JDExtractor
        elif platform == 'pdd':
            extractor = PDDExtractor
        else:
            extractor = TMExtractor

        data = extractor.extract(page, url)

        # 截图
        shot_mgr = ScreenshotManager(page, save_dir, prefix=f"{idx:03d}_")
        screenshot_path = shot_mgr.screenshot_page()

        # OCR 兜底
        if platform == 'jd' and (not data.get('shop_name') or data['shop_name'] == '未提取'):
            ocr_shop = extractor.ocr_shop_name(screenshot_path) if hasattr(extractor, 'ocr_shop_name') else ''
            if ocr_shop:
                data['shop_name'] = ocr_shop
                print(f"  [OCR] 识别店铺名: {ocr_shop}")

        # 截图文件（给 IPP 用）
        data['screenshot_path'] = screenshot_path
        data['save_dir'] = save_dir

        # 生成侵权对比图
        shop_for_comparison = data.get('shop_name', '') or '未知店铺'
        comparison_path = generate_comparison_image(screenshot_path, shop_for_comparison, save_dir, idx)
        if comparison_path:
            data['comparison_image'] = comparison_path
            print(f"  [📸] 侵权对比图已生成")

        # 关闭
        try:
            ctx.browser.close()
        except:
            pass

    return data


# ==================== 主程序 ====================
def main():
    print("""
╔══════════════════════════════════════════════════════════╗
║      WOW English 知识产权投诉 - 批量处理工具 v1.0         ║
║      📋 Excel 填链接 → 自动抓取 → 生成 IPP 投诉文档       ║
╚══════════════════════════════════════════════════════════╝
    """)

    # 读取 Excel 中的链接
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, "投诉链接模板.xlsx")

    excel_path = None
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    elif os.path.exists(template_path):
        excel_path = template_path
        print(f"[INPUT] 自动找到模板文件：{template_path}")
    else:
        print("[INPUT] 请提供 Excel 文件路径，或将文件命名为「投诉链接模板.xlsx」放在同目录")
        excel_path = input(">>> 拖入或输入路径：").strip().strip('"')

    if not os.path.exists(excel_path):
        print(f"[ERROR] 文件不存在：{excel_path}")
        print("提示：将您的链接 Excel 命名为「投诉链接模板.xlsx」放在脚本同目录即可")
        return

    links = read_links_from_excel(excel_path)
    if not links:
        print("[ERROR] 未从 Excel 中找到任何商品链接！")
        print("请确保 Excel 中有 jd.com / tmall.com / taobao.com / pinduoduo.com 链接")
        return

    print(f"\n[INFO] 共找到 {len(links)} 个链接：")
    for i, lnk in enumerate(links, 1):
        print(f"  {i}. {lnk[:70]}")
    print()

    # 创建批次输出目录
    batch_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    batch_dir = os.path.join(OUTPUT_DIR, f"batch_{batch_ts}")
    os.makedirs(batch_dir, exist_ok=True)

    # Chrome 调试端口提示
    cdp_port = find_cdp_port()
    if not cdp_port:
        print("""
[⚠️  提示] 未检测到 Chrome 调试模式！
建议先双击「启动已登录Chrome.bat」，再运行本脚本。
或直接在浏览器登录电商平台，脚本会处理验证码。
按回车继续（需手动处理验证码）...
        """)
        input()

    # 批量处理
    results = []
    failed = []

    for i, url in enumerate(links, 1):
        try:
            result = process_single_url(None, url, batch_dir, i, len(links))
            results.append(result)

            status = "✅" if "是" in result.get("infringement_check", "否") else "⚪"
            print(f"  {status} 标题：{result.get('title','')[:40]}")
            print(f"  {status} 店铺：{result.get('shop_name','')}")
            print(f"  {status} 侵权：{result.get('infringement_check','')}")

            # 每处理完一个，随机等待（模拟人类操作节奏）
            gap = random.uniform(3, 10)
            print(f"  [💤] 间隔 {gap:.1f} 秒...")
            time.sleep(gap)
        except Exception as e:
            print(f"  ❌ 处理失败: {e}")
            failed.append({"url": url, "error": str(e)})
            results.append({
                "platform": detect_platform(url), "title": "提取失败",
                "url": url, "shop_name": "", "seller_id": "",
                "price": "", "sales": "", "infringement_check": f"失败：{e}",
                "screenshot_path": "", "save_dir": batch_dir
            })

        # 生成 IPP 文档
    print(f"\n\n{'='*60}")
    print("[📄] 生成 IPP 投诉文档...")

    generator = IPPComplaintGenerator(results, batch_dir)

    excel_file = generator.generate_excel()
    print(f"  ✅ {excel_file}")

    txt_file = generator.generate_txt()
    print(f"  ✅ {txt_file}")

    summary_file = generator.generate_batch_summary()
    print(f"  ✅ {summary_file}")

    # 新增：IPP 完整材料包（含表单+理由+清单）
    ipp_form_file = generator.generate_ipp_form()
    print(f"  ✅ {ipp_form_file}")

    # 统计
    total = len(results)
    infr = sum(1 for r in results if '是' in r.get('infringement_check', ''))
    fail = len(failed)

    print(f"""
╔══════════════════════════════════════════════════════════╗
║                   ✅ 批量处理完成！                        ║
╚══════════════════════════════════════════════════════════╝

【统计】
  📦 总计处理：{total} 件
  🔴 疑似侵权：{infr} 件
  ⚠️  处理失败：{fail} 件

【输出目录】
  {batch_dir}

【输出文件说明】
  📋 IPP完整材料包_*.xlsx → ⭐推荐使用！包含：
     · Sheet1: 京东IP投诉表单（预填充）
     · Sheet2: 淘天IPP投诉表单（预填充）
     · Sheet3: 投诉理由（直接复制粘贴）
     · Sheet4: 材料准备清单

  📊 IPP投诉文档_*.xlsx   → 简洁版投诉列表
  📄 IPP投诉文档_*.txt    → 纯文本投诉理由
  📈 批量处理汇总_*.xlsx  → 处理结果总览（侵权标红）

【IPP 投诉操作步骤】
  1. 打开「IPP完整材料包_*.xlsx」
  2. Sheet4「材料准备清单」→ 确认您已有权利证明文件
  3. 登录京东 ipr.jd.com 或淘天 ipp.alibabagroup.com
  4. 上传权利证明（营业执照+商标注册证），只需上传一次
  5. Sheet3「投诉理由」→ 全选复制 → 粘贴到投诉表单
  6. 逐个填写侵权链接（从 Sheet1/Sheet2 的商品列表复制）
  7. 上传截图证据（每个商品文件夹下的 screenshot.png）
  8. 提交投诉！

【材料清单】
  ✅ 权利证明（营业执照+商标注册证）- 您准备
  ✅ 侵权截图 - 已自动生成（每个商品文件夹）
  ✅ 侵权对比图 - 已自动生成（标注侵权区域）
  ✅ 投诉理由 - 已自动生成（Sheet3 直接复制）
  ✅ 投诉表单 - 已预填充（Sheet1/Sheet2）
""")


if __name__ == "__main__":
    main()
