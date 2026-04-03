"""
WOW English 京东维权提取器 v2.0
专门针对京东平台的增强版提取工具
"""

import os
import sys
import io
import time
import random
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

# Windows 编码修复
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    try:
        os.system('chcp 65001 >nul 2>&1')
    except:
        pass

try:
    from playwright.sync_api import sync_playwright, Page
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    print(f"[ERROR] Missing dependency: {e}")
    print("Run: pip install playwright openpyxl")
    sys.exit(1)


# ==================== 配置 ====================
BRAND_KEYWORDS = [
    "wow english", "wowenglish",
    "史蒂夫英语", "Steve English",
    "英语启蒙动画", "English Singsing"
]

OUTPUT_DIR = Path("output")


# ==================== 京东提取器 ====================
class JDExtractor:
    """京东商品信息提取器 - 增强版"""

    # 京东常用选择器（2024年新版页面）
    TITLE_SELECTORS = [
        # 新版页面
        '[class*="product-title"]',
        '[class*="sku-name"]',
        '.p-title',
        '#itemName',
        # 旧版兼容
        '.itemInfo-wrap h1',
        'div[class*="product-name"] h1',
        '#name h1',
        'h1[class*="name"]',
        # 兜底
        'h1'
    ]

    PRICE_SELECTORS = [
        # 新版页面
        '[class*="price"] [class*="integer"]',
        '.price J-p-',
        '.p-price .price',
        '#jdprice',
        '[data-price]',
        # 旧版兼容
        '.itemInfo-wrap .p-price',
        '[class*="sale-price"]',
        # 兜底
        '[class*="price-wrap"] span'
    ]

    SHOP_SELECTORS = [
        # 新版页面
        '[class*="shop-name"] a',
        '.shop-name a',
        '#pop_shop .name',
        '[class*="seller"] a',
        # 旧版兼容
        '.itemInfo-wrap .p-shop a',
        '.p-shop a',
        '[class*="store"] a',
        # 兜底
        '[class*="shop"]'
    ]

    SELLER_ID_SELECTORS = [
        '#pop_shop .name',
        '[class*="seller-code"]',
        '[class*="shop-id"]',
        '[data-sellerid]',
        '.shop-name'
    ]

    SALES_SELECTORS = [
        # 新版页面
        '[class*="sale"]',
        '[class*="count"]',
        '.p-commit a',
        '#commit_count',
        # 旧版兼容
        '.itemInfo-wrap .p-commit',
        '[class*="evaluate"]',
        '[class*="comment"]'
    ]

    MAIN_IMG_SELECTORS = [
        '#spec-img',
        '[class*="spec-img"] img',
        '[class*="main-img"] img',
        '[class*="preview"] img',
        '#large',
        '[class*="viewer"] img'
    ]

    @classmethod
    def extract(cls, page: Page, url: str) -> dict:
        """提取京东商品信息"""
        data = {
            "platform": "京东",
            "title": "",
            "price": "",
            "shop_name": "",
            "seller_id": "",
            "sales": "",
            "main_images": [],
            "url": url,
            "infringement_check": "否"
        }

        # 1. 商品标题
        data["title"] = cls._try_selectors(page, cls.TITLE_SELECTORS)

        # 2. 价格
        data["price"] = cls._try_selectors(page, cls.PRICE_SELECTORS)

        # 3. 店铺名称
        data["shop_name"] = cls._try_selectors(page, cls.SHOP_SELECTORS)

        # 4. 卖家ID
        data["seller_id"] = cls._try_selectors(page, cls.SELLER_ID_SELECTORS)

        # 5. 销量/评价数
        data["sales"] = cls._try_selectors(page, cls.SALES_SELECTORS)

        # 6. 主图
        data["main_images"] = cls._extract_main_images(page)

        # 7. 侵权检查
        data["infringement_check"] = cls._check_infringement(data)

        return data

    @staticmethod
    def _try_selectors(page: Page, selectors: list) -> str:
        """尝试多个选择器，返回第一个有效结果"""
        for sel in selectors:
            try:
                el = page.locator(sel).first
                if el.count() > 0 and el.is_visible():
                    text = el.inner_text().strip()
                    if text and len(text) > 2:
                        return text[:200]  # 截断超长文本
            except:
                continue
        return "未提取"

    @staticmethod
    def _extract_main_images(page: Page) -> list:
        """提取主图URL"""
        images = []
        for sel in JDExtractor.MAIN_IMG_SELECTORS:
            try:
                els = page.locator(sel).all()
                for el in els[:5]:  # 最多5张
                    if el.is_visible():
                        src = el.get_attribute('src') or el.get_attribute('data-src')
                        if src and 'blank' not in src.lower():
                            # 京东图片URL处理：去掉缩略图后缀
                            src = re.sub(r'/\d+x\d+/', '/800x800/', src)
                            images.append(src)
                if images:
                    break
            except:
                continue
        return list(set(images))  # 去重

    @staticmethod
    def _check_infringement(data: dict) -> str:
        """检查侵权关键词"""
        text = f"{data.get('title', '')} {data.get('shop_name', '')}".lower()
        matched = [kw for kw in BRAND_KEYWORDS if kw.lower() in text]
        return f"是（匹配：{', '.join(matched)}）" if matched else "否"


# ==================== 截图管理器 ====================
class ScreenshotManager:
    """页面截图和图片下载"""

    def __init__(self, page: Page, save_dir: Path):
        self.page = page
        self.save_dir = save_dir
        self.save_dir.mkdir(parents=True, exist_ok=True)
        (self.save_dir / "main_images").mkdir(exist_ok=True)

    def screenshot_page(self, full_page: bool = True) -> str:
        """截取页面截图"""
        filename = f"{self.save_dir}/screenshot.png"

        if full_page:
            # 全页截图
            page_height = self.page.evaluate("document.body.scrollHeight")
            self.page.set_viewport_size({"width": 1280, "height": min(page_height, 5000)})

        self.page.screenshot(path=filename, full_page=full_page)
        return filename

    def download_main_images(self, urls: list) -> list:
        """下载商品主图"""
        saved = []
        for i, url in enumerate(urls[:5]):
            try:
                # 清理URL
                url = url.replace('_50x50', '').replace('_b.jpg', '.jpg')
                filename = f"{self.save_dir}/main_images/jd_main_{i+1}.jpg"

                response = self.page.context.request.get(url)
                if response.ok:
                    with open(filename, 'wb') as f:
                        f.write(response.body())
                    saved.append(filename)
            except Exception as e:
                print(f"  [WARN] Image download failed: {e}")
        return saved


# ==================== Excel 导出 ====================
class ExcelExporter:
    """导出到Excel"""

    def __init__(self, data: dict, save_dir: Path):
        self.data = data
        self.save_dir = save_dir

    def export(self) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "京东维权信息"

        # 表头样式
        header_fill = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")  # 京东红
        header_font = Font(bold=True, color="FFFFFF", size=11)

        headers = ["平台", "商品标题", "商品链接", "价格", "店铺名称",
                   "卖家ID", "销量/评价", "主图URL", "提取时间", "疑似侵权"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")

        # 数据行
        main_urls = "\n".join(self.data.get("main_images", []))
        row_data = [
            self.data.get("platform", ""),
            self.data.get("title", ""),
            self.data.get("url", ""),
            self.data.get("price", ""),
            self.data.get("shop_name", ""),
            self.data.get("seller_id", ""),
            self.data.get("sales", ""),
            main_urls,
            self.data.get("timestamp", ""),
            self.data.get("infringement_check", "")
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=2, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")

        # 列宽
        widths = [10, 45, 55, 12, 25, 20, 15, 60, 20, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # 侵权标红
        if "是" in self.data.get("infringement_check", ""):
            ws.cell(row=2, column=10).fill = PatternFill(start_color="FF0000", fill_type="solid")
            ws.cell(row=2, column=10).font = Font(bold=True, color="FFFFFF")

        filename = f"{self.save_dir}/jd_extraction.xlsx"
        wb.save(filename)
        return filename


# ==================== 举报模板 ====================
class ReportGenerator:
    """生成举报模板"""

    def __init__(self, data: dict, save_dir: Path):
        self.data = data
        self.save_dir = save_dir

    def generate(self) -> str:
        template = f"""
{'='*60}
               京东平台侵权举报信息
{'='*60}

【商品信息】
平台：{self.data.get('platform')}
商品标题：{self.data.get('title')}
商品链接：{self.data.get('url')}
当前价格：{self.data.get('price')}

【店铺信息】
店铺名称：{self.data.get('shop_name')}
卖家ID：{self.data.get('seller_id')}
销量/评价：{self.data.get('sales')}

【侵权分析】
关键词匹配：{self.data.get('infringement_check')}

【证据文件】
- 页面截图：screenshot.png
- 商品主图：main_images/
- 结构化数据：jd_extraction.xlsx

{'='*60}
提取时间：{self.data.get('timestamp')}
证据目录：{self.save_dir}
{'='*60}
"""
        filename = f"{self.save_dir}/jd_report.txt"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(template)
        return filename


# ==================== 主程序 ====================
def main():
    print("""
╔══════════════════════════════════════════════════════════╗
║      WOW English 京东维权提取工具 v2.0                   ║
╚══════════════════════════════════════════════════════════╝
    """)

    # 获取链接
    url = None
    if len(sys.argv) > 1:
        url = sys.argv[1]
    elif len(sys.argv) > 2 and sys.argv[1] == '--jd':
        url = sys.argv[2]
    else:
        # 尝试从stdin读取
        print("[INPUT] 请输入京东商品链接:")
        url = input(">>> ").strip()

    if not url or 'jd.com' not in url:
        print("[ERROR] 请提供京东商品链接")
        print("[USAGE] python jd_extract.py <商品链接>")
        return

    # 创建输出目录
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_dir = OUTPUT_DIR / f"jd_{timestamp}"
    save_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n[1/4] URL: {url}")
    print(f"[2/4] Output: {save_dir}")

    # Cookie 存储路径
    cookie_path = Path("jd_cookies.json")

    # 首先尝试接管已打开的 Chrome 浏览器
    debug_port = None
    try:
        import socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        # 尝试常见的 CDP 调试端口
        for port in [9222, 9223, 9224]:
            result = sock.connect_ex(('127.0.0.1', port))
            if result == 0:
                debug_port = port
                break
        sock.close()
    except:
        pass

    print("[3/4] Connecting to browser...")
    with sync_playwright() as p:
        browser = None
        page = None

        # ── 方案A：接管已打开的 Chrome（需要 --remote-debugging-port）──
        if debug_port:
            try:
                browser = p.chromium.connect_over_cdp(f"http://127.0.0.1:{debug_port}")
                # 找到已打开的京东标签页，或新开一个标签页
                default_context = browser.contexts[0]
                jd_pages = [pg for pg in default_context.pages if 'jd.com' in pg.url]
                if jd_pages:
                    page = jd_pages[0]
                    print(f"[CDP] 已接管现有京东标签页: {page.url[:60]}")
                else:
                    page = default_context.new_page()
                    print("[CDP] 在已有浏览器中新开标签页")
            except Exception as e:
                print(f"[CDP] 接管失败: {e}，改用独立浏览器...")
                browser = None

        # ── 方案B：启动新浏览器（加载 Cookie）──
        if browser is None:
            import json
            browser = p.chromium.launch(
                headless=False,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--disable-dev-shm-usage',
                    '--no-sandbox',
                    '--start-maximized'
                ]
            )
            context = browser.new_context(
                viewport={"width": 1366, "height": 768},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                locale='zh-CN',
                no_viewport=True
            )
            # 加载已保存的 Cookie
            if cookie_path.exists():
                with open(cookie_path, 'r', encoding='utf-8') as f:
                    cookies = json.load(f)
                context.add_cookies(cookies)
                print(f"[COOKIE] 已加载上次登录的 Cookie")
            else:
                print("[COOKIE] 未找到 Cookie，如需要请在弹出浏览器中登录")
            page = context.new_page()

        # 反爬：移除 webdriver 特征
        try:
            page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
                window.chrome = { runtime: {} };
            """)
        except:
            pass  # CDP 模式下 add_init_script 可能不可用，忽略

        try:
            # 访问目标商品页
            print(f"[4/4] Loading page...")
            page.goto(url, wait_until="domcontentloaded", timeout=30000)

            # 检查是否被重定向到登录页
            current_url = page.url
            if "login" in current_url or "passport" in current_url:
                print("\n" + "="*55)
                print("  [LOGIN] 请在浏览器中登录京东，完成后脚本自动继续")
                print("="*55)
                try:
                    page.wait_for_url("**item.jd.com**", timeout=120000)
                    print("[LOGIN] 登录成功！")
                    page.goto(url, wait_until="domcontentloaded", timeout=30000)
                except Exception:
                    print("[LOGIN] 超时，尝试继续提取...")

            # 保存 Cookie（方案B才保存，方案A已是用户真实 Chrome 不需要）
            if not debug_port:
                import json
                cookies = page.context.cookies()
                with open(cookie_path, 'w', encoding='utf-8') as f:
                    json.dump(cookies, f, ensure_ascii=False, indent=2)
                print(f"[COOKIE] 已保存 Cookie，下次免登录")

            # 等待内容加载
            page.wait_for_timeout(3000)

            # 滚动触发懒加载
            for i in range(3):
                page.evaluate(f"window.scrollTo(0, {i * 500})")
                page.wait_for_timeout(500)

            # 提取数据
            print("\n[EXTRACTING] Extracting product info...")
            extractor = JDExtractor()
            data = extractor.extract(page, url)
            data["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 截图
            print("[SCREENSHOT] Taking screenshot...")
            shot_mgr = ScreenshotManager(page, save_dir)
            shot_mgr.screenshot_page()

            # 下载主图
            if data.get("main_images"):
                print(f"[IMAGES] Downloading {len(data['main_images'])} main images...")
                shot_mgr.download_main_images(data["main_images"])

            # 导出Excel
            print("[EXCEL] Exporting data...")
            exporter = ExcelExporter(data, save_dir)
            exporter.export()

            # 生成报告
            print("[REPORT] Generating report...")
            ReportGenerator(data, save_dir).generate()

            # 输出结果
            print(f"""
╔══════════════════════════════════════════════════════════╗
║                    EXTRACTION COMPLETE                   ║
╚══════════════════════════════════════════════════════════╝

[RESULT]
  Platform: {data['platform']}
  Title: {data['title'][:60]}...
  Price: {data['price']}
  Shop: {data['shop_name']}
  Infringement: {data['infringement_check']}

[OUTPUT]
  {save_dir}

[FILES]
  jd_extraction.xlsx - Excel data
  jd_report.txt      - Report template
  screenshot.png     - Page screenshot
  main_images/       - Product images
            """)

            # 写入日志
            with open(OUTPUT_DIR / "jd_log.txt", "a", encoding="utf-8") as f:
                f.write(f"[{data['timestamp']}] {data['platform']} | {data['title'][:40]} | {url}\n")

        except Exception as e:
            print(f"\n[ERROR] Extraction failed: {e}")
            print("\nTroubleshooting:")
            print("  1. Check if URL is valid")
            print("  2. Login to JD if required")
            print("  3. Try again or use headless=False to debug")
        finally:
            browser.close()


if __name__ == "__main__":
    main()
