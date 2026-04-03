"""
WOW English 电商维权信息提取工具
专为版权方设计的半自动提取工具，支持淘宝/天猫、京东、拼多多
"""

import os
import sys
import io
import json
import base64
import asyncio
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

# Windows 编码修复
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    # 设置控制台代码页
    try:
        os.system('chcp 65001 >nul 2>&1')
    except:
        pass

try:
    from playwright.sync_api import sync_playwright, Page, Browser
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    print(f"❌ 缺少依赖包: {e}")
    print("请运行: pip install playwright openpyxl")
    sys.exit(1)


# ==================== 配置 ====================
BRAND_KEYWORDS = [
    "wow english", "Wow English", "WOW English",
    "Wow english", "wowenglish",
    "史蒂夫英语", "Steve English", "steve english",
    "英语启蒙动画", "English Singsing", "singsing"
]

OUTPUT_DIR = Path("output")
PLATFORMS = {
    "taobao": ["taobao.com", "tmall.com", "tm.jd.com"],
    "jd": ["jd.com", "jd.hk", "mall.jd.com"],
    "pinduoduo": ["yangkeduo.com", "pinduoduo.com"]
}


# ==================== 平台检测 ====================
def detect_platform(url: str) -> str:
    """检测商品所属平台"""
    parsed = urlparse(url.lower())
    domain = parsed.netloc

    for platform, domains in PLATFORMS.items():
        if any(d in domain for d in domains):
            if platform == "taobao":
                return "淘宝/天猫"
            elif platform == "jd":
                return "京东"
            elif platform == "pinduoduo":
                return "拼多多"
    return "未知平台"


# ==================== 信息提取器 ====================
class ProductExtractor:
    """各平台商品信息提取器"""

    @staticmethod
    def extract_taobao(page: Page) -> dict:
        """提取淘宝/天猫商品信息"""
        data = {
            "platform": "淘宝/天猫",
            "title": "",
            "price": "",
            "shop_name": "",
            "seller_id": "",
            "sales": "",
            "main_images": [],
            "url": page.url
        }

        # 商品标题
        selectors = [
            'h1[class*="title"]',
            '.tb-title h3',
            '[class*="product-name"]',
            '#J_DetailTitle .tb-title'
        ]
        for sel in selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible():
                    data["title"] = el.inner_text().strip()
                    break
            except:
                pass

        # 价格
        price_selectors = [
            '.tb-rmb-num',
            '[class*="price"] span',
            '.price-wrapper .price',
            '#J_SetPrice .tb-price'
        ]
        for sel in price_selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible():
                    data["price"] = el.inner_text().strip()
                    break
            except:
                pass

        # 店铺名称
        shop_selectors = [
            '.shop-name a',
            '.shopName a',
            '[class*="shop-name"] a'
        ]
        for sel in shop_selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible():
                    data["shop_name"] = el.inner_text().strip()
                    break
            except:
                pass

        # 卖家ID
        seller_selectors = [
            '#J_PShop .shopAddress span:first-child',
            '[class*="shop-info"] span:first-child'
        ]
        for sel in seller_selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible():
                    data["seller_id"] = el.inner_text().strip()
                    break
            except:
                pass

        # 销量
        sales_selectors = [
            '#J_DetailCounter .J_DetailCounter',
            '[class*="sale"]',
            '.tb-sell-counter'
        ]
        for sel in sales_selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible():
                    data["sales"] = el.inner_text().strip()
                    break
            except:
                pass

        # 主图
        try:
            main_img = page.locator('#J_ImgBooth, .tb-main-pic img, [class*="main-pic"] img').first
            if main_img.is_visible():
                src = main_img.get_attribute('src') or main_img.get_attribute('data-src')
                if src and 'blank' not in src.lower():
                    data["main_images"].append(src)
        except:
            pass

        return data

    @staticmethod
    def extract_jd(page: Page) -> dict:
        """提取京东商品信息"""
        data = {
            "platform": "京东",
            "title": "",
            "price": "",
            "shop_name": "",
            "seller_id": "",
            "sales": "",
            "main_images": [],
            "url": page.url
        }

        # 商品标题
        selectors = ['.product-title__china', '#name .product-title', 'h1[itemprop="name"]']
        for sel in selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible():
                    data["title"] = el.inner_text().strip()
                    break
            except:
                pass

        # 价格
        price_selectors = ['.price J-p-', '.itemprice', '#jdprice']
        for sel in price_selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible():
                    data["price"] = el.inner_text().strip()
                    break
            except:
                pass

        # 店铺名称
        try:
            shop = page.locator('[class*="shopName"] a, .shop-name a, #pop_shop .name').first
            if shop.is_visible():
                data["shop_name"] = shop.inner_text().strip()
        except:
            pass

        # 销量
        try:
            sales = page.locator('[class*="sale"] span, .stock-value').first
            if sales.is_visible():
                data["sales"] = sales.inner_text().strip()
        except:
            pass

        # 主图
        try:
            main_img = page.locator('#spec-img, .product-preview img, [class*="main-pic"] img').first
            if main_img.is_visible():
                src = main_img.get_attribute('src') or main_img.get_attribute('data-src')
                if src:
                    data["main_images"].append(src)
        except:
            pass

        return data

    @staticmethod
    def extract_pinduoduo(page: Page) -> dict:
        """提取拼多多商品信息"""
        data = {
            "platform": "拼多多",
            "title": "",
            "price": "",
            "shop_name": "",
            "seller_id": "",
            "sales": "",
            "main_images": [],
            "url": page.url
        }

        # 商品标题
        selectors = ['.goods-title, [class*="title"]', '.product-name', 'h1']
        for sel in selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible():
                    text = el.inner_text().strip()
                    if text and len(text) > 5:
                        data["title"] = text
                        break
            except:
                pass

        # 价格
        price_selectors = ['.price, [class*="price"] .amount, .goods-price']
        for sel in price_selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible():
                    data["price"] = el.inner_text().strip()
                    break
            except:
                pass

        # 店铺名称
        try:
            shop = page.locator('[class*="shop-name"] a, .store-name').first
            if shop.is_visible():
                data["shop_name"] = shop.inner_text().strip()
        except:
            pass

        # 主图
        try:
            main_img = page.locator('[class*="main-pic"] img, .goods-img img').first
            if main_img.is_visible():
                src = main_img.get_attribute('src') or main_img.get_attribute('data-src')
                if src:
                    data["main_images"].append(src)
        except:
            pass

        return data

    @staticmethod
    def extract(page: Page) -> dict:
        """根据URL自动选择提取方法"""
        platform = detect_platform(page.url)

        if "淘宝" in platform or "天猫" in platform:
            return ProductExtractor.extract_taobao(page)
        elif "京东" in platform:
            return ProductExtractor.extract_jd(page)
        elif "拼多多" in platform:
            return ProductExtractor.extract_pinduoduo(page)
        else:
            # 通用提取
            return {
                "platform": platform,
                "title": page.title(),
                "price": "未提取",
                "shop_name": "未提取",
                "seller_id": "未提取",
                "sales": "未提取",
                "main_images": [],
                "url": page.url
            }


# ==================== 截图下载器 ====================
class ScreenshotManager:
    """页面截图和图片下载"""

    def __init__(self, page: Page, save_dir: Path):
        self.page = page
        self.save_dir = save_dir
        self.save_dir.mkdir(parents=True, exist_ok=True)
        (self.save_dir / "main_images").mkdir(exist_ok=True)

    def screenshot_page(self) -> str:
        """截取整页截图"""
        filename = f"{self.save_dir}/screenshot.png"

        # 全屏截图
        page_height = self.page.evaluate("document.body.scrollHeight")
        self.page.set_viewport_size({"width": 1280, "height": page_height})

        self.page.screenshot(path=filename, full_page=True)
        print(f"✅ 页面截图已保存: {filename}")
        return filename

    def download_main_images(self, urls: list) -> list:
        """下载商品主图"""
        saved_paths = []
        for i, url in enumerate(urls[:5]):  # 最多下载5张
            try:
                # 处理URL
                url = url.replace('_50x50', '').replace('_b.jpg', '.jpg')

                filename = f"{self.save_dir}/main_images/image_{i+1}.jpg"
                self.page.context.request.get(url)
                self.page.context.request.get(url)

                # 使用内容请求下载
                response = self.page.context.request.get(url)
                if response.ok:
                    with open(filename, 'wb') as f:
                        f.write(response.body())
                    saved_paths.append(filename)
                    print(f"✅ 主图 {i+1} 已保存")
            except Exception as e:
                print(f"⚠️ 下载图片失败: {e}")

        return saved_paths


# ==================== Excel 导出 ====================
class ExcelExporter:
    """导出提取数据到Excel"""

    def __init__(self, data: dict, save_dir: Path):
        self.data = data
        self.save_dir = save_dir
        self.save_dir.mkdir(parents=True, exist_ok=True)

    def check_infringement(self) -> str:
        """检查是否包含侵权关键词"""
        text = f"{self.data.get('title', '')}".lower()
        matched = [kw for kw in BRAND_KEYWORDS if kw.lower() in text]
        return "是" if matched else "否"

    def export(self) -> str:
        """导出到Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "维权信息"

        # 标题样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        headers = ["平台", "商品标题", "商品链接", "价格", "店铺名称",
                   "卖家ID", "销量", "主图URL", "提取时间", "疑似侵权"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 写入数据
        main_urls = "\n".join(self.data.get("main_images", []))
        infringement = self.check_infringement()

        values = [
            self.data.get("platform", ""),
            self.data.get("title", ""),
            self.data.get("url", ""),
            self.data.get("price", ""),
            self.data.get("shop_name", ""),
            self.data.get("seller_id", ""),
            self.data.get("sales", ""),
            main_urls,
            self.data.get("timestamp", ""),
            infringement
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # 设置列宽
        column_widths = [12, 40, 50, 12, 20, 20, 12, 60, 20, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # 高亮侵权商品
        if infringement == "是":
            cell = ws.cell(row=2, column=10)
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        filename = f"{self.save_dir}/extraction_data.xlsx"
        wb.save(filename)
        print(f"✅ Excel 已保存: {filename}")
        return filename


# ==================== 举报模板生成 ====================
class ReportGenerator:
    """生成举报模板"""

    def __init__(self, data: dict, save_dir: Path):
        self.data = data
        self.save_dir = save_dir

    def generate(self) -> str:
        """生成举报模板"""
        infringement = "是" if "是" in self.data.get("infringement_check", "") else "否"

        template = f"""
{'='*60}
                    电商平台侵权举报模板
{'='*60}

【基本信息】
平台：{self.data.get('platform', '')}
商品标题：{self.data.get('title', '')}
商品链接：{self.data.get('url', '')}

【卖家信息】
店铺名称：{self.data.get('shop_name', '')}
卖家ID：{self.data.get('seller_id', '')}
当前价格：{self.data.get('price', '')}
销量：{self.data.get('sales', '')}

【侵权判定】
是否涉及 WOW English 关键词：{infringement}

【侵权事实描述】
（请根据实际情况填写）
本店销售的商品未经 WOW English 版权方授权，擅自使用品牌名称
和相关内容，涉嫌构成商标侵权及不正当竞争。

【证据说明】
1. 商品页面截图：见同目录 screenshot.png
2. 商品主图：见同目录 main_images/ 文件夹
3. 相关页面信息已保存至 extraction_data.xlsx

【举报依据】
1. 涉嫌侵犯注册商标专用权
2. 涉嫌构成不正当竞争
3. 未经授权使用他人品牌名称和标识

【处理请求】
1. 要求平台立即下架该商品
2. 要求对侵权店铺进行处罚
3. 保留进一步追究法律责任的权利

【联系方式】
（请填写版权方联系方式）

{'='*60}
模板生成时间：{self.data.get('timestamp', '')}
证据保存位置：{self.save_dir}
{'='*60}
"""

        filename = f"{self.save_dir}/report_template.txt"
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(template)

        print(f"✅ 举报模板已生成: {filename}")
        return filename


# ==================== 主程序 ====================
def main():
    print("""
╔══════════════════════════════════════════════════════════════╗
║     WOW English 电商维权信息提取工具 v1.0                      ║
║     专为版权方设计的半自动提取工具                            ║
╚══════════════════════════════════════════════════════════════╝
    """)

    # 解析参数
    url = None
    if len(sys.argv) > 1:
        url = sys.argv[1]
    else:
        # 尝试从剪贴板获取
        try:
            import pyperclip
            clipboard_content = pyperclip.paste()
            if clipboard_content and ('taobao' in clipboard_content or
                                       'tmall' in clipboard_content or
                                       'jd.com' in clipboard_content or
                                       'yangkeduo' in clipboard_content or
                                       'pinduoduo' in clipboard_content):
                url = clipboard_content
                print(f"📋 从剪贴板检测到商品链接: {url[:50]}...")
        except ImportError:
            print("💡 提示：安装 pyperclip 可自动从剪贴板读取链接")
            print("   pip install pyperclip")

    if not url:
        print("❌ 未检测到商品链接")
        print("\n使用方法：")
        print("  python extract_product.py <商品链接>")
        print("\n或：")
        print("  1. 复制商品链接到剪贴板")
        print("  2. 运行 python extract_product.py（需安装pyperclip）")
        return

    # 创建输出目录
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    save_dir = OUTPUT_DIR / timestamp
    save_dir.mkdir(parents=True, exist_ok=True)

    # 启动浏览器
    print("\n🚀 启动浏览器...")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            viewport={"width": 1280, "height": 800},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )
        page = context.new_page()

        try:
            # 访问商品页
            print(f"\n📱 正在访问: {url}")
            page.goto(url, wait_until="networkidle", timeout=30000)

            # 等待页面加载
            print("⏳ 等待页面渲染...")
            page.wait_for_timeout(2000)

            # 提取信息
            print("\n📊 正在提取商品信息...")
            extractor = ProductExtractor()
            data = extractor.extract(page)
            data["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            data["platform"] = detect_platform(url)

            # 截图
            print("\n📸 正在截图...")
            screenshot_mgr = ScreenshotManager(page, save_dir)
            screenshot_mgr.screenshot_page()

            # 下载主图
            if data.get("main_images"):
                screenshot_mgr.download_main_images(data["main_images"])

            # 导出 Excel
            print("\n📁 正在导出数据...")
            excel_exporter = ExcelExporter(data, save_dir)
            excel_exporter.export()

            # 生成举报模板
            print("\n📝 正在生成举报模板...")
            report_gen = ReportGenerator(data, save_dir)
            report_gen.generate()

            # 输出结果
            print("\n" + "="*60)
            print("✅ 提取完成！")
            print("="*60)
            print(f"""
【提取结果】

📦 平台：{data.get('platform')}
📝 标题：{data.get('title', '未提取')[:50]}...
💰 价格：{data.get('price', '未提取')}
🏪 店铺：{data.get('shop_name', '未提取')}
📊 销量：{data.get('sales', '未提取')}

【保存位置】
📁 {save_dir}

【包含文件】
  📄 extraction_data.xlsx  - 结构化数据（维权用）
  📄 report_template.txt   - 举报模板（直接复制使用）
  📸 screenshot.png         - 页面截图
  📁 main_images/           - 商品主图

【下一步】
  1. 打开 extraction_data.xlsx 核对信息
  2. 复制 report_template.txt 内容到平台举报表单
  3. 将截图和主图作为证据附件上传
            """)

            # 保存日志
            log_file = OUTPUT_DIR / "log.txt"
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(f"[{data['timestamp']}] {data['platform']} | {data.get('title', '')[:30]} | {url}\n")

        except Exception as e:
            print(f"\n❌ 提取失败: {e}")
            print("\n可能原因：")
            print("  1. 网络连接问题")
            print("  2. 页面加载超时")
            print("  3. 该页面需要登录")
            print("\n请尝试：")
            print("  1. 确认商品链接可正常访问")
            print("  2. 在浏览器中手动打开页面后重试")
            print("  3. 尝试使用 headless=False 模式查看问题")

        finally:
            browser.close()


if __name__ == "__main__":
    main()
